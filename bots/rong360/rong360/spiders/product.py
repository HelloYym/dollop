# -*- coding: utf-8 -*-
from __future__ import print_function
from __future__ import unicode_literals
import scrapy
from utils.webpage import get_trunk, get_content
from utils.exporter import read_cache
from rong360.items import ProductItem
import json
from datetime import datetime
import copy
from openpyxl import load_workbook
import pkgutil


class ProductSpider(scrapy.Spider):
    name = "product"
    allowed_domains = ["rong360.com"]

    pipeline = ['UniqueItemPersistencePipeline']

    start_urls = ['https://www.rong360.com/licai-net/']

    def __init__(self, name=None, **kwargs):
        super(ProductSpider, self).__init__(name, **kwargs)

        self.product_company_dict = json.loads(pkgutil.get_data('rong360', 'resources/baobao.json'))

        # self.product_company_dict = json.load(open('resources/baobao.json'))
        # self.product_company_dict = dict()
        # wb = load_workbook('resources/baobao.xlsx')
        # ws = wb.active
        # for (name, issuer, company) in ws.iter_rows(min_row=2, max_col=3, max_row=73):
        #     # print(name.value, issuer.value, company.value)
        #     self.product_company_dict[name.value] = company.value.strip()


    def get_code_from_url(self, url):
        return url.split('-')[-1]

    def parse(self, response):
        for product_abs in response.xpath('//tbody[@id="ui_product_list_tbody"]/tr'):
            product_href = product_abs.xpath('@click-url').extract_first()

            name = get_content(product_abs.xpath('td[1]//text()').extract())
            issuer = get_content(product_abs.xpath('td[2]//text()').extract())
            fund_scale = get_content(product_abs.xpath('td[3]//text()').extract())
            w_income = get_content(product_abs.xpath('td[4]//text()').extract())
            rate = get_content(product_abs.xpath('td[5]//text()').extract())
            min_amount = get_content(product_abs.xpath('td[6]//text()').extract())
            ceiling = get_content(product_abs.xpath('td[7]//text()').extract())
            speed = get_content(product_abs.xpath('td[8]//text()').extract())
            company = self.product_company_dict[name]

            yield scrapy.FormRequest(url=product_href,
                                     meta={'name': name, 'issuer': issuer, 'company': company, 'fund_scale': fund_scale,
                                           'w_income': w_income,
                                           'rate': rate, 'min_amount': min_amount, 'ceiling': ceiling, 'speed': speed},
                                     callback=self.parse_product_detail,
                                     dont_filter=True)

    def parse_product_detail(self, response):
        product = ProductItem()

        product['code'] = self.get_code_from_url(response.url)
        product['link'] = response.url
        product['name'] = response.meta['name']
        product['issuer'] = response.meta['issuer']
        product['company'] = response.meta['company']
        product['fund_scale'] = response.meta['fund_scale']
        product['w_income'] = response.meta['w_income']
        product['min_amount'] = response.meta['min_amount']
        product['ceiling'] = response.meta['ceiling']
        product['speed'] = response.meta['speed']

        product['rate'] = response.meta['rate']

        product['ceiling_comment'] = get_content(response.xpath('//div[@id="blockLeft"]/text()').extract())
        product['speed_comment'] = get_content(response.xpath('//div[@id="blockRight"]/text()').extract())

        if (response.xpath('//p').re_first(r'合作机构(.*),')):
            product['cooperation'] = response.xpath('//p').re_first(r'合作机构(.*),').strip()

        seq = json.loads(response.xpath('//script').re_first(r'history_7_charts = (.*)}') + '}')['data']

        for (timestamp, rate) in seq[-1:]:
            product_day = copy.deepcopy(product)
            product_day['date'] = datetime.fromtimestamp(timestamp / 1000).strftime('%Y-%m-%d')
            product_day['rate'] = rate
            yield product_day
            # print(product_day)
